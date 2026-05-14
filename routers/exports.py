from __future__ import annotations

import csv
import io
from datetime import date

from fastapi import APIRouter, Depends
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session

from auth import get_current_user
from database import get_db
from models import Investment, Scenario, User
from services.pdf_report import generate_pdf

router = APIRouter(prefix="/exports", tags=["exports"])


@router.get("/csv")
async def export_portfolio_csv(current: User = Depends(get_current_user), db: Session = Depends(get_db)) -> StreamingResponse:
    rows = db.query(Investment).filter(Investment.user_id == current.id).all()
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["id", "name", "type", "symbol", "amount_invested", "current_value", "purchase_date", "notes", "roi_pct"])
    for r in rows:
        roi = ((r.current_value - r.amount_invested) / r.amount_invested * 100.0) if r.amount_invested > 0 else 0.0
        writer.writerow([r.id, r.name, r.type, r.symbol or "", r.amount_invested, r.current_value, r.purchase_date.isoformat(), r.notes or "", round(roi, 2)])
    buf.seek(0)
    filename = f"portfolio-{date.today().isoformat()}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/scenarios.csv")
async def export_scenarios_csv(current: User = Depends(get_current_user), db: Session = Depends(get_db)) -> StreamingResponse:
    rows = db.query(Scenario).filter(Scenario.user_id == current.id).all()
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["id", "name", "amount", "horizon_months", "annual_return_pct", "inflation_rate_pct", "risk_level"])
    for r in rows:
        writer.writerow([r.id, r.name, r.amount, r.horizon_months, r.annual_return, r.inflation_rate, r.risk_level])
    buf.seek(0)
    filename = f"scenarios-{date.today().isoformat()}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/pdf")
async def export_pdf(current: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Response:
    pdf = generate_pdf(db, current)
    filename = f"investment-report-{date.today().isoformat()}.pdf"
    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/xlsx")
async def export_xlsx(current: User = Depends(get_current_user), db: Session = Depends(get_db)) -> Response:
    """Multi-sheet Excel workbook with positions, scenarios and a summary."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    investments = db.query(Investment).filter(Investment.user_id == current.id).all()
    scenarios = db.query(Scenario).filter(Scenario.user_id == current.id).all()

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8A7558", end_color="8A7558", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center")

    def _style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=12)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Sheet 1: Positions
    ws = wb.active
    ws.title = "Positions"
    ws.append(["Name", "Type", "Symbol", "Quantity", "Amount Invested (USD)",
               "Current Value (USD)", "ROI %", "Purchase Date", "Account", "Notes"])
    for r in investments:
        roi = ((r.current_value - r.amount_invested) / r.amount_invested * 100) if r.amount_invested else 0
        ws.append([
            r.name, r.type, r.symbol or "", r.quantity or "",
            r.amount_invested, r.current_value, round(roi, 2),
            r.purchase_date.isoformat() if r.purchase_date else "",
            r.account_type or "", r.notes or "",
        ])
    _style_header(ws)

    # Sheet 2: Real estate detail (only for property rows)
    re_rows = [r for r in investments if r.type == "real_estate"]
    if re_rows:
        ws_re = wb.create_sheet("Real estate")
        ws_re.append(["Name", "Address", "City", "Postal code", "Country",
                      "Surface (m²)", "Garden (m²)", "Rent/month (USD)",
                      "Charges/month (USD)", "Loan amount (USD)",
                      "Mortgage/month (USD)", "Loan rate %"])
        for r in re_rows:
            ws_re.append([
                r.name, r.address or "", r.city or "", r.postal_code or "",
                r.country or "", r.surface_sqm or "", r.garden_sqm or "",
                r.monthly_rental_income or "", r.monthly_rental_charges or "",
                r.loan_amount or "", r.monthly_mortgage_payment or "",
                r.loan_interest_rate_pct or "",
            ])
        _style_header(ws_re)

    # Sheet 3: Scenarios
    if scenarios:
        ws_sc = wb.create_sheet("Scenarios")
        ws_sc.append(["Name", "Amount (USD)", "Horizon (months)",
                      "Annual return %", "Inflation %", "Risk level"])
        for s in scenarios:
            ws_sc.append([s.name, s.amount, s.horizon_months,
                          s.annual_return, s.inflation_rate, s.risk_level])
        _style_header(ws_sc)

    # Sheet 4: Summary
    ws_sum = wb.create_sheet("Summary")
    total_inv = sum(r.amount_invested for r in investments)
    total_cur = sum(r.current_value for r in investments)
    total_roi = ((total_cur - total_inv) / total_inv * 100) if total_inv else 0
    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["Generated", date.today().isoformat()])
    ws_sum.append(["Currency", current.currency])
    ws_sum.append(["Number of positions", len(investments)])
    ws_sum.append(["Total invested (USD)", round(total_inv, 2)])
    ws_sum.append(["Current value (USD)", round(total_cur, 2)])
    ws_sum.append(["Total ROI %", round(total_roi, 2)])
    _style_header(ws_sum)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"portfolio-{date.today().isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
